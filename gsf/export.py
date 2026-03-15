from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.styles import PatternFill, Font, Border, Side

from gsf.constants import (
    CONF_EXCEL_COLORS,
    CONF_HIGH,
    CONF_MODERATE,
    CONF_VERY_HIGH,
    GEO_CLOSE_KM,
    GEO_MODERATE_KM,
    GEO_VERY_CLOSE_KM,
    MODE_ORDER,
    AITCH_STEP,
    EXCEL_COLORS,
    EXCEL_OVERFLOW,
    aitch_col,
)


def _get_fill_for_value(
    val: Any,  # noqa: ANN401
    mode_key: str = "trace",
) -> PatternFill | None:
    """Return a paired-color PatternFill using fixed step per mode."""
    try:
        v = float(val)
    except (ValueError, TypeError):
        return None
    step = AITCH_STEP.get(mode_key, 0.8)
    hex_color = EXCEL_OVERFLOW
    for i, c in enumerate(EXCEL_COLORS, start=1):
        if v < i * step:
            hex_color = c
            break
    return PatternFill(
        start_color=hex_color, fill_type="solid",
    )


def _get_conf_fill(val: Any) -> PatternFill | None:  # noqa: ANN401
    """Return a PatternFill for a Conf percentage value (4 bands)."""
    try:
        v = float(val)
    except (ValueError, TypeError):
        return None
    if v > CONF_VERY_HIGH:
        c = CONF_EXCEL_COLORS["very_high"]
    elif v >= CONF_HIGH:
        c = CONF_EXCEL_COLORS["high"]
    elif v >= CONF_MODERATE:
        c = CONF_EXCEL_COLORS["moderate"]
    else:
        c = CONF_EXCEL_COLORS["low"]
    return PatternFill(start_color=c, fill_type="solid")


def _get_geo_fill(val: Any) -> PatternFill | None:  # noqa: ANN401
    """Return a PatternFill for a Geo Dist string value."""
    try:
        num = float(str(val).replace(" km", ""))
    except (ValueError, TypeError, AttributeError):
        return None
    if num < GEO_VERY_CLOSE_KM:
        c = EXCEL_COLORS[0]
    elif num < GEO_CLOSE_KM:
        c = EXCEL_COLORS[1]
    elif num < GEO_MODERATE_KM:
        c = EXCEL_COLORS[2]
    else:
        c = EXCEL_OVERFLOW
    return PatternFill(start_color=c, fill_type="solid")


def create_styled_excel(
    results_df: pd.DataFrame,
    query_accession: str,
    query_site: str,
    direction: str,
    enabled_modes: list[str] | None = None,
) -> bytes:
    """Create a styled Excel workbook with color-coded cells."""
    if enabled_modes is None:
        enabled_modes = list(MODE_ORDER)

    output = BytesIO()
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:  # pyright: ignore[reportArgumentType]
        results_df.to_excel(
            writer, sheet_name="Results",
            index=True, startrow=2,
        )
        ws = writer.sheets["Results"]

        # Header metadata
        ws["A1"] = (
            f"Query: {query_accession} — {query_site}"
        )
        ws["A1"].font = Font(bold=True, size=12, name="Georgia")
        dir_label = (
            "Artefact -> Geology"
            if direction == "artefact_to_geology"
            else "Geology -> Artefact"
        )
        ws["A2"] = (
            f"Direction: {dir_label} | "
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )

        # Find column positions for distance columns
        header_row = 3
        col_map: dict[str, int] = {}
        aitch_col_names = {
            aitch_col(m) for m in enabled_modes
        }
        special_cols = {"Geo Dist", "Conf"}
        for cell in ws[header_row]:
            if (
                cell.value in aitch_col_names
                or cell.value in special_cols
            ):
                col_map[cell.value] = cell.column

        # Build mode_key lookup per Aitch column
        aitch_mode_map: dict[int, str] = {}
        for mode in enabled_modes:
            col_name = aitch_col(mode)
            if col_name in col_map:
                aitch_mode_map[col_map[col_name]] = mode

        geo_col_num = col_map.get("Geo Dist")
        conf_col_num = col_map.get("Conf")

        # Apply fills to data rows
        for row in ws.iter_rows(
            min_row=header_row + 1, max_row=ws.max_row,
        ):
            for cell in row:
                cell.border = thin_border
                if cell.column in aitch_mode_map:
                    fill = _get_fill_for_value(
                        cell.value,
                        aitch_mode_map[cell.column],
                    )
                    if fill:
                        cell.fill = fill
                elif cell.column == conf_col_num:
                    fill = _get_conf_fill(cell.value)
                    if fill:
                        cell.fill = fill
                elif cell.column == geo_col_num:
                    fill = _get_geo_fill(cell.value)
                    if fill:
                        cell.fill = fill

        # Style header row
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.border = thin_border

    return output.getvalue()
