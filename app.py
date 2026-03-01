from __future__ import annotations

import os
import json
import base64
import time
import logging
from datetime import datetime
from io import BytesIO
from collections import Counter
from functools import partial
from typing import Any
from urllib.parse import quote

import requests
import pandas as pd
import numpy as np
from geopy.distance import geodesic
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from openpyxl.styles import PatternFill, Font, Border, Side

# -----------------------------
# Page Configuration (must be first Streamlit call)
# -----------------------------
st.set_page_config(
    page_title="GSF - Geology Source Finder",
    page_icon="\U0001FAA8",
    layout="wide",
    initial_sidebar_state="expanded",
)

# -----------------------------
# Logging
# -----------------------------
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# -----------------------------
# Constants
# -----------------------------
EPSILON = 1e-10

# Metadata columns that precede ratio data in the Excel sheets
METADATA_COLUMNS = {
    "n", "Accession #", "Site", "Region", "Type", "Lithology",
}

# Aitchison distance thresholds for color coding
AITCH_VERY_STRONG = 1.0
AITCH_STRONG = 2.0
AITCH_MODERATE = 3.0

# Euclidean distance thresholds for color coding
EUCL_VERY_STRONG = 2.0
EUCL_STRONG = 4.0
EUCL_MODERATE = 6.0

# ALR-5 configuration
ALR5_ELEMENTS = ["Ni", "Cr", "Y", "Nb", "Sr", "Zr"]
ALR5_DENOMINATOR = "Zr"
ALR5_NUMERATORS = ["Ni", "Cr", "Y", "Nb", "Sr"]
ALR5_RATIO_NAMES = [
    "ln(Ni/Zr)", "ln(Cr/Zr)", "ln(Y/Zr)",
    "ln(Nb/Zr)", "ln(Sr/Zr)",
]
ALR5_D = 6  # number of parts in the sub-composition

# ALR-5 Aitchison thresholds (scaled by sqrt(5)/sqrt(16))
ALR5_AITCH_VERY_STRONG = 0.56
ALR5_AITCH_STRONG = 1.12
ALR5_AITCH_MODERATE = 1.68

# Geographical distance thresholds (km) for color coding
GEO_VERY_CLOSE_KM = 60
GEO_CLOSE_KM = 120
GEO_MODERATE_KM = 200

# Distance rounding precision for display
GEO_DIST_ROUND_TO_KM = 10

# Default / max number of results
DEFAULT_TOP_N = 20
MAX_TOP_N = 100

# Color scheme for threshold visualization
COLORS = {
    "very_strong": "#ADD8E6",  # Light blue
    "strong": "#90EE90",       # Light green
    "moderate": "#FFDAB9",     # Peach
    "weak": "#F08080",         # Light coral
}

CELL_STYLE = "color: black; border: 1px solid gray"

# Lithology color map for scatter plots
LITHOLOGY_COLOR_MAP: dict[str, str] = {
    "Ophite-THOL": "#FF0000",
    "Ophite-ALK": "#8B0000",
    "Amphibolite": "#0000FF",
    "Eclogite": "#EE82EE",
    "Metabasite": "#00FF00",
    "Basalt-ALK": "#000000",
    "Gabbro-THOL": "#808080",
    "Gabbro-CALC": "#D3D3D3",
    "Gabbro-ALK": "#696969",
    "Metagabbro": "#90EE90",
}

# Region color map for scatter plots
REGION_COLOR_MAP: dict[str, str] = {
    "Upper Guadalquivir": "orange",
    "Southeast": "dodgerblue",
    "Subbaetic": "red",
    "Middle Guadalquivir": "limegreen",
    "Lower Guadalquivir": "mediumpurple",
}


def _aitch_thresholds(
    mode_key: str,
) -> tuple[float, float, float]:
    """Return Aitchison thresholds for the given mode."""
    if mode_key == "alr5":
        return (
            ALR5_AITCH_VERY_STRONG,
            ALR5_AITCH_STRONG,
            ALR5_AITCH_MODERATE,
        )
    return (AITCH_VERY_STRONG, AITCH_STRONG, AITCH_MODERATE)


def _has_euclidean(mode_key: str) -> bool:
    """Return True if the mode includes Euclidean distance."""
    return mode_key != "alr5"


# -----------------------------
# Microsoft Graph / OneDrive Constants
# -----------------------------
_GRAPH_AUTHORITY = "https://login.microsoftonline.com/common"
_GRAPH_DEVICE_CODE_URL = f"{_GRAPH_AUTHORITY}/oauth2/v2.0/devicecode"
_GRAPH_TOKEN_URL = f"{_GRAPH_AUTHORITY}/oauth2/v2.0/token"
_GRAPH_SCOPES = "Files.Read offline_access"
_GRAPH_BASE = "https://graph.microsoft.com/v1.0"


# -----------------------------
# Data Loading
# -----------------------------
def _load_alr5_df(
    filepath: str, sheet_name: str,
) -> pd.DataFrame:
    """Load ppm data and compute ALR-5 coordinates.

    Reads the raw ppm sheet, extracts 6 elements
    (Ni, Cr, Y, Nb, Sr, Zr) and computes ln(X/Zr)
    for each numerator. Missing or non-positive values
    are left as NaN so that incomplete ALR coordinates
    are excluded from distance calculations rather than
    being replaced with extreme outliers.
    """
    raw = pd.read_excel(filepath, sheet_name=sheet_name)

    meta_cols = [
        c for c in METADATA_COLUMNS if c in raw.columns
    ]

    missing = [
        e for e in ALR5_ELEMENTS if e not in raw.columns
    ]
    if missing:
        raise ValueError(
            f"Missing elements {missing} in sheet "
            f"'{sheet_name}' of {filepath}"
        )

    result = raw[meta_cols].copy()
    elem_vals: dict[str, pd.Series] = {}
    for elem in ALR5_ELEMENTS:
        vals = pd.to_numeric(raw[elem], errors="coerce")
        # Non-positive values → NaN; avoids log-space
        # outliers (ln(ε/x) ≈ -25) that dominate distances
        vals = vals.where(vals > 0, np.nan)
        elem_vals[elem] = vals

    zr = elem_vals[ALR5_DENOMINATOR]
    for num, name in zip(
        ALR5_NUMERATORS, ALR5_RATIO_NAMES,
    ):
        # NaN in numerator or denominator propagates to NaN
        result[name] = np.log(elem_vals[num] / zr)

    result["Accession #"] = (
        result["Accession #"].astype(str)
    )
    return result


@st.cache_data(show_spinner=False)
def load_data(
    element_mode: str = "trace",
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load required Excel files from a 'Data' folder.

    Args:
        element_mode: "trace" for 16 ratio columns,
            "all" for 22 ratio columns,
            "alr5" for 5 ALR log-ratio coordinates.

    Returns:
        (artefact_df, geology_df, geo_coords_df, arch_coords_df)
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    possible_dirs = [
        os.path.join(base_dir, "Data"),
        os.path.join(base_dir, "..", "Data"),
    ]

    if element_mode == "all":
        filenames = {
            "artefact": "AXEs metabasite data (All Elements).xlsx",
            "geology": "Geology samples data (All Elements).xlsx",
            "coords": "Coordinates sheet.xlsx",
        }
    else:
        # Both "trace" and "alr5" use the Trace Elements files
        filenames = {
            "artefact": "AXEs metabasite data (Trace Elements).xlsx",
            "geology": "Geology samples data (Trace Elements).xlsx",
            "coords": "Coordinates sheet.xlsx",
        }

    found_data_dir = None
    for d in possible_dirs:
        paths = [os.path.join(d, f) for f in filenames.values()]
        if all(os.path.exists(p) for p in paths):
            found_data_dir = d
            logger.info("Data found in directory: %s", d)
            break

    if not found_data_dir:
        raise FileNotFoundError(
            f"Required files not found in: {possible_dirs}\n"
            "Place the Excel files in a 'Data' folder "
            "next to this script."
        )

    try:
        if element_mode == "alr5":
            artefact_df = _load_alr5_df(
                os.path.join(
                    found_data_dir, filenames["artefact"],
                ),
                "AXEs ppm data",
            )
            geology_df = _load_alr5_df(
                os.path.join(
                    found_data_dir, filenames["geology"],
                ),
                "Geology ppm data",
            )
        else:
            artefact_df = pd.read_excel(
                os.path.join(
                    found_data_dir, filenames["artefact"],
                ),
                sheet_name="AXEs Ratios",
            )
            geology_df = pd.read_excel(
                os.path.join(
                    found_data_dir, filenames["geology"],
                ),
                sheet_name="Geology ratios",
            )
        coords_sheets = pd.read_excel(
            os.path.join(found_data_dir, filenames["coords"]),
            sheet_name=[
                "Geology Samples Coord",
                "Archaeology Sites Coord",
            ],
        )
        geo_coords_df = coords_sheets["Geology Samples Coord"]
        arch_coords_df = coords_sheets["Archaeology Sites Coord"]
    except Exception as e:
        logger.exception("Error loading Excel files.")
        st.error(
            "Error loading Excel files. "
            "Check file names and sheet names."
        )
        raise e

    artefact_df["Accession #"] = (
        artefact_df["Accession #"].astype(str)
    )
    geology_df["Accession #"] = (
        geology_df["Accession #"].astype(str)
    )
    geo_coords_df["Accession #"] = (
        geo_coords_df["Accession #"].astype(str)
    )

    # Longitude values in the coordinate sheets are inconsistently stored:
    # most Spanish sites use positive values (e.g. 3.47 instead of -3.47)
    # while some sites (e.g. Vila Nova de São Pedro) are already negative.
    # All sites are in the Iberian Peninsula (west of the prime meridian),
    # so force all longitudes to negative with -abs() to normalise both cases.
    geo_coords_df["Longitude"] = geo_coords_df["Longitude"].abs() * -1
    arch_coords_df["Longitude"] = arch_coords_df["Longitude"].abs() * -1

    return artefact_df, geology_df, geo_coords_df, arch_coords_df


@st.cache_data(show_spinner=False)
def _load_photo_mapping() -> pd.DataFrame:
    """Load artefact-to-photo filename mapping from CSV."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    for d in [
        os.path.join(base_dir, "Data"),
        os.path.join(base_dir, "..", "Data"),
    ]:
        path = os.path.join(d, "photo_mapping.csv")
        if os.path.exists(path):
            df = pd.read_csv(path, encoding="utf-8")
            df["Accession #"] = df["Accession #"].astype(str).str.strip()
            return df
    return pd.DataFrame()


def _get_config_path() -> str:
    """Return path to the app config JSON file."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, "Data", "app_config.json")


def _load_config() -> dict:
    """Load persisted app configuration."""
    path = _get_config_path()
    if os.path.exists(path):
        try:
            with open(path, encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            pass
    return {}


def _save_config(config: dict) -> None:
    """Save app configuration to JSON file."""
    path = _get_config_path()
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=2)
    except OSError:
        logger.warning("Could not save config to %s", path)


# -----------------------------
# OAuth Token Cache
# -----------------------------
def _get_token_cache_path() -> str:
    """Return path to the OAuth token cache JSON file."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, "Data", "token_cache.json")


def _load_token_cache() -> dict:
    """Load cached OAuth tokens from disk."""
    path = _get_token_cache_path()
    if os.path.exists(path):
        try:
            with open(path, encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            pass
    return {}


def _save_token_cache(data: dict) -> None:
    """Save OAuth tokens to disk."""
    path = _get_token_cache_path()
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
    except OSError:
        logger.warning("Could not save token cache to %s", path)


# -----------------------------
# OAuth2 Device Code Flow
# -----------------------------
def _start_device_code_flow(client_id: str) -> dict | None:
    """Request a device code from Microsoft identity platform.

    Returns dict with device_code, user_code, verification_uri,
    or None on failure.
    """
    try:
        resp = requests.post(
            _GRAPH_DEVICE_CODE_URL,
            data={
                "client_id": client_id,
                "scope": _GRAPH_SCOPES,
            },
            timeout=15,
        )
        resp.raise_for_status()
        return resp.json()
    except Exception as exc:
        logger.error("Device code request failed: %s", exc)
        return None


def _poll_for_token(
    client_id: str, device_code: str,
) -> dict | None:
    """Single non-blocking poll to exchange device code for tokens.

    Returns token dict on success, None if still pending.
    Raises ValueError if the flow expired or was denied.
    """
    try:
        resp = requests.post(
            _GRAPH_TOKEN_URL,
            data={
                "client_id": client_id,
                "grant_type": (
                    "urn:ietf:params:oauth:grant-type:device_code"
                ),
                "device_code": device_code,
            },
            timeout=15,
        )
        data = resp.json()

        if "access_token" in data:
            return data

        error = data.get("error", "")
        if error in ("authorization_pending", "slow_down"):
            return None

        raise ValueError(
            data.get("error_description", f"Auth failed: {error}")
        )
    except requests.RequestException as exc:
        logger.error("Token poll failed: %s", exc)
        return None


def _refresh_access_token(
    client_id: str, refresh_token: str,
) -> dict | None:
    """Use refresh token to get a new access token."""
    try:
        resp = requests.post(
            _GRAPH_TOKEN_URL,
            data={
                "client_id": client_id,
                "grant_type": "refresh_token",
                "refresh_token": refresh_token,
                "scope": _GRAPH_SCOPES,
            },
            timeout=15,
        )
        resp.raise_for_status()
        return resp.json()
    except Exception as exc:
        logger.error("Token refresh failed: %s", exc)
        return None


def _get_valid_access_token(client_id: str) -> str | None:
    """Return a valid Graph API access token, refreshing if needed.

    Checks session_state → disk cache → refresh token.
    Returns None if not authenticated.
    """
    token_data = st.session_state.get("graph_token_data")

    if not token_data:
        token_data = _load_token_cache()
        if token_data and "access_token" in token_data:
            st.session_state["graph_token_data"] = token_data
        else:
            return None

    # Check if token is still valid (5-min buffer)
    expires_at = token_data.get("expires_at", 0)
    if time.time() < expires_at - 300:
        return token_data.get("access_token")

    # Token expired — try refresh
    refresh_token = token_data.get("refresh_token")
    if not refresh_token:
        return None

    new_data = _refresh_access_token(client_id, refresh_token)
    if not new_data:
        st.session_state.pop("graph_token_data", None)
        return None

    new_data["expires_at"] = (
        time.time() + new_data.get("expires_in", 3600)
    )
    st.session_state["graph_token_data"] = new_data
    _save_token_cache(new_data)
    return new_data["access_token"]


# -----------------------------
# Graph API Helpers
# -----------------------------
def _encode_share_url(share_url: str) -> str:
    """Encode a OneDrive share URL into a Graph API sharing token.

    Format: 'u!' + base64url(share_url) with no padding.
    """
    encoded = base64.urlsafe_b64encode(
        share_url.encode("utf-8"),
    ).rstrip(b"=").decode("ascii")
    return f"u!{encoded}"


def _resolve_graph_path(relative_path: str) -> str:
    """Apply CSV-to-disk path corrections for Graph API."""
    for csv_prefix, disk_prefix in _PATH_PREFIXES.items():
        if relative_path.startswith(csv_prefix):
            return disk_prefix + relative_path[len(csv_prefix):]
    return relative_path


def _resolve_share_folder(
    access_token: str, share_url: str,
) -> tuple[str, str] | None:
    """Resolve a share URL to (drive_id, folder_path).

    Caches result in session_state.  Returns None on failure.
    """
    cache_key = "_share_resolved"
    cached = st.session_state.get(cache_key)
    if cached:
        return cached

    share_token = _encode_share_url(share_url)
    url = (
        f"{_GRAPH_BASE}/shares/{share_token}/driveItem"
    )
    try:
        resp = requests.get(
            url,
            headers={
                "Authorization": f"Bearer {access_token}",
            },
            params={
                "$select": "id,name,parentReference",
            },
            timeout=15,
        )
        if not resp.ok:
            try:
                err_body = resp.json()
            except Exception:
                err_body = resp.text[:200]
            st.session_state["_graph_last_error"] = (
                f"Share resolve HTTP {resp.status_code}: "
                f"{err_body}"
            )
            return None

        item = resp.json()
        drive_id = item["parentReference"]["driveId"]
        parent_path = item["parentReference"].get(
            "path", "",
        )
        # parent_path: "/drives/{id}/root:/some/path"
        if "/root:" in parent_path:
            base = parent_path.split(
                "/root:",
            )[1].lstrip("/")
        else:
            base = ""
        folder_name = item.get("name", "")
        folder_path = (
            f"{base}/{folder_name}" if base
            else folder_name
        )
        result = (drive_id, folder_path)
        st.session_state[cache_key] = result
        return result
    except Exception as exc:
        st.session_state["_graph_last_error"] = str(exc)
        return None



def _fetch_image_urls_from_graph(
    access_token: str,
    share_url: str,
    relative_path: str,
) -> tuple[str, str] | None:
    """Get pre-authenticated URLs for an image via Graph API.

    Returns (thumbnail_url, full_res_url) or None on failure.
    The browser loads images directly from Microsoft's CDN —
    no server-side download or base64 encoding needed.
    """
    resolved = _resolve_share_folder(
        access_token, share_url,
    )
    if not resolved:
        return None
    drive_id, folder_path = resolved

    full_path = f"{folder_path}/{relative_path}"
    encoded_path = quote(full_path, safe="/")
    url = (
        f"{_GRAPH_BASE}/drives/{drive_id}"
        f"/root:/{encoded_path}:"
    )
    try:
        resp = requests.get(
            url,
            headers={
                "Authorization": f"Bearer {access_token}",
            },
            params={"$expand": "thumbnails"},
            timeout=15,
        )
        if not resp.ok:
            try:
                err_body = resp.json()
            except Exception:
                err_body = resp.text[:200]
            st.session_state["_graph_last_error"] = (
                f"HTTP {resp.status_code} for "
                f"{full_path}: {err_body}"
            )
            return None

        data = resp.json()
        full_url = data.get(
            "@microsoft.graph.downloadUrl", "",
        )
        if not full_url:
            st.session_state["_graph_last_error"] = (
                f"No download URL for {full_path}"
            )
            return None

        thumb_url = full_url
        thumbs = data.get("thumbnails", [])
        if thumbs and "large" in thumbs[0]:
            thumb_url = thumbs[0]["large"].get(
                "url", full_url,
            )
        return (thumb_url, full_url)
    except Exception as exc:
        st.session_state["_graph_last_error"] = str(exc)
        return None


_IMAGE_URL_CACHE_TTL = 1800  # 30 min (URLs expire ~1 hr)


def _get_cached_image_urls(
    access_token: str,
    share_url: str,
    relative_path: str,
) -> tuple[str, str] | None:
    """Fetch image URLs with session-level caching."""
    cache_key = f"imgurl_{relative_path}"
    cached = st.session_state.get(cache_key)
    if cached is not None:
        cached_time, cached_urls = cached
        if time.time() - cached_time < _IMAGE_URL_CACHE_TTL:
            return cached_urls

    urls = _fetch_image_urls_from_graph(
        access_token, share_url, relative_path,
    )
    if urls:
        st.session_state[cache_key] = (time.time(), urls)
    return urls


# Path corrections: CSV relative paths → actual disk paths
_PATH_PREFIXES: dict[str, str] = {
    "Montefrío/": "UGR assemblages/Montefrío/",
    "Los Millares/": "UGR assemblages/Los Millares/",
    "El Malagón/": "UGR assemblages/El Malagón/",
    "Museo Almería/": "Museo Almería pictures/",
}


def _resolve_photo_path(
    base_dir: str, relative_path: str,
) -> str | None:
    """Resolve a CSV relative path to an actual file on disk."""
    # Apply known prefix corrections
    resolved = relative_path
    for csv_prefix, disk_prefix in _PATH_PREFIXES.items():
        if relative_path.startswith(csv_prefix):
            resolved = disk_prefix + relative_path[len(csv_prefix):]
            break

    full_path = os.path.join(base_dir, resolved)
    if os.path.isfile(full_path):
        return full_path

    # Fallback: try original path unchanged
    fallback = os.path.join(base_dir, relative_path)
    if os.path.isfile(fallback):
        return fallback

    return None


def _get_artefact_image_paths(
    accession: str,
    photo_df: pd.DataFrame,
    photos_base_dir: str,
) -> list[str]:
    """Return local file paths for an accession's photos."""
    if photo_df.empty or not photos_base_dir:
        return []
    row = photo_df[
        photo_df["Accession #"] == str(accession).strip()
    ]
    if row.empty:
        return []
    row = row.iloc[0]
    paths = []
    for col in [
        "Image_1", "Image_2", "Image_3",
        "Image_4", "Image_5", "Image_6",
    ]:
        val = row.get(col)
        if pd.notna(val) and str(val).strip():
            resolved = _resolve_photo_path(
                photos_base_dir, str(val).strip(),
            )
            if resolved:
                paths.append(resolved)
    return paths


def _get_artefact_images(
    accession: str,
    photo_df: pd.DataFrame,
    photos_base_dir: str = "",
    access_token: str = "",
    share_url: str = "",
) -> list[str | bytes | tuple[str, str]]:
    """Return image sources for an accession's photos.

    Tries local files first (if photos_base_dir is set),
    falls back to Graph API (if access_token and share_url).
    Returns list of:
    - str: local file path
    - bytes: image data
    - tuple[str, str]: (thumbnail_url, full_res_url)
    """
    if photo_df.empty:
        return []
    row = photo_df[
        photo_df["Accession #"] == str(accession).strip()
    ]
    if row.empty:
        return []
    row = row.iloc[0]

    images: list[str | bytes | tuple[str, str]] = []
    for col in [
        "Image_1", "Image_2", "Image_3",
        "Image_4", "Image_5", "Image_6",
    ]:
        val = row.get(col)
        if not (pd.notna(val) and str(val).strip()):
            continue
        relative = str(val).strip()

        # Try local path first
        if photos_base_dir:
            local = _resolve_photo_path(
                photos_base_dir, relative,
            )
            if local:
                images.append(local)
                continue

        # Try Graph API (URL-based, no download)
        if access_token and share_url:
            graph_path = _resolve_graph_path(relative)
            urls = _get_cached_image_urls(
                access_token, share_url, graph_path,
            )
            if urls:
                images.append(urls)

    return images


# -----------------------------
# Ratio Column Extraction
# -----------------------------
def _get_ratio_columns(sample: pd.Series) -> list[str]:
    """Extract ratio column names, excluding metadata and NaN."""
    ratio_data = sample.drop(
        labels=[c for c in METADATA_COLUMNS if c in sample.index],
    )
    return list(ratio_data.dropna().index)


def _count_total_ratio_columns(df: pd.DataFrame) -> int:
    """Count total ratio columns (all non-metadata columns)."""
    return len(
        [c for c in df.columns if c not in METADATA_COLUMNS]
    )


# -----------------------------
# Vectorized Distance Calculations
# -----------------------------
def _compute_distances_vectorized(
    query_ratios: np.ndarray, target_matrix: np.ndarray,
) -> tuple[np.ndarray, np.ndarray]:
    """Compute Aitchison and Euclidean distances.

    Uses numpy broadcasting between one query vector
    and a matrix of target vectors.

    The Aitchison distance uses the Centered Log-Ratio transform:
        CLR(x) = log(x) - mean(log(x))
        d_A(x, y) = ||CLR(x) - CLR(y)||_2
    """
    q = np.where(
        np.isfinite(query_ratios) & (query_ratios > 0),
        query_ratios, EPSILON,
    )
    t = np.where(
        np.isfinite(target_matrix) & (target_matrix > 0),
        target_matrix, EPSILON,
    )

    log_q = np.log(q)
    clr_q = log_q - np.mean(log_q)

    log_t = np.log(t)
    clr_t = log_t - np.mean(log_t, axis=1, keepdims=True)

    aitchison_dists = np.linalg.norm(clr_t - clr_q, axis=1)
    euclidean_dists = np.linalg.norm(
        target_matrix - query_ratios, axis=1,
    )

    return aitchison_dists, euclidean_dists


def _compute_alr5_aitchison_vectorized(
    query_alr: np.ndarray, target_matrix: np.ndarray,
) -> np.ndarray:
    """Compute Aitchison distance from ALR coordinates.

    Formula: d_A^2 = sum(d_i^2) - (sum(d_i))^2 / D
    where d_i = alr_i(x) - alr_i(y),
    D = k + 1 (k ALR coordinates + Zr reference element).

    D is inferred from input shape so partial compositions
    (when some elements are missing) are handled correctly.
    ALR values are already log-ratios; no log transform needed.
    """
    k = query_alr.shape[0]   # number of ALR coordinates
    D = k + 1                # k numerators + Zr denominator
    diffs = target_matrix - query_alr
    sum_sq = np.sum(diffs ** 2, axis=1)
    sq_sum = np.sum(diffs, axis=1) ** 2
    d_a_sq = np.maximum(sum_sq - sq_sum / D, 0.0)
    return np.sqrt(d_a_sq)


def _compute_geodesic(
    origin: tuple[float, float] | None,
    destination: tuple[float, float] | None,
) -> str:
    """Compute geodesic distance in km, formatted as a string."""
    if origin is None or destination is None:
        return "Unknown"
    try:
        dist_km = geodesic(origin, destination).km
        rounded = int(
            round(dist_km / GEO_DIST_ROUND_TO_KM)
            * GEO_DIST_ROUND_TO_KM
        )
        return f"{rounded} km"
    except (ValueError, TypeError):
        return "Unknown"


# -----------------------------
# Unified Match Function
# -----------------------------
def get_top_matches(
    query_sample: pd.Series,
    target_df: pd.DataFrame,
    query_coords_df: pd.DataFrame,
    target_coords_df: pd.DataFrame,
    top_n: int = DEFAULT_TOP_N,
    direction: str = "artefact_to_geology",
    mode_key: str = "trace",
) -> pd.DataFrame:
    """Find the top-N matching samples by compositional similarity.

    Computes Aitchison and Euclidean distances (vectorized),
    sorts by [Aitchison, Euclidean], takes top_n, then computes
    geodesic distances only for those results.
    """
    ratio_cols = _get_ratio_columns(query_sample)
    if not ratio_cols:
        return pd.DataFrame()

    if direction == "artefact_to_geology":
        keep_cols = [
            "Lithology", "Accession #", "Site", "Region",
        ]
    else:
        keep_cols = ["Accession #", "Site", "Region"]

    available_cols = [
        c for c in keep_cols + ratio_cols
        if c in target_df.columns
    ]
    aligned_df = target_df[available_cols].dropna(  # pyright: ignore[reportCallIssue]
        subset=ratio_cols,
    )

    if aligned_df.empty:
        return pd.DataFrame()

    query_series = pd.Series(pd.to_numeric(
        query_sample[ratio_cols], errors="coerce",
    )).fillna(EPSILON)
    query_vector = np.asarray(query_series, dtype=float)
    target_matrix = np.asarray(
        aligned_df[ratio_cols]
        .apply(pd.to_numeric, errors="coerce")
        .fillna(EPSILON),
        dtype=float,
    )

    results_df = aligned_df[keep_cols].copy()
    results_df = results_df.reset_index(drop=True)

    if mode_key == "alr5":
        aitch_dists = _compute_alr5_aitchison_vectorized(
            query_vector, target_matrix,
        )
        results_df["Aitch Dist"] = np.round(aitch_dists, 2)
        results_df = results_df.sort_values(
            by=["Aitch Dist"],
        ).head(top_n)
    else:
        aitch_dists, eucl_dists = (
            _compute_distances_vectorized(
                query_vector, target_matrix,
            )
        )
        results_df["Aitch Dist"] = np.round(aitch_dists, 2)
        results_df["Eucl Dist"] = np.round(eucl_dists, 2)
        results_df = results_df.sort_values(  # pyright: ignore[reportCallIssue]
            by=["Aitch Dist", "Eucl Dist"],
        ).head(top_n)
    results_df.reset_index(drop=True, inplace=True)
    results_df.index += 1

    # Compute geodesic distances only for top_n results
    if direction == "artefact_to_geology":
        site_match = query_coords_df[
            query_coords_df["Site"] == query_sample["Site"]
        ]
        query_coords: tuple[float, float] | None = (
            (
                float(site_match.iloc[0]["Latitude"]),
                float(site_match.iloc[0]["Longitude"]),
            )
            if not site_match.empty
            else None
        )
        target_coord_lookup: dict[
            str, tuple[float, float]
        ] = {}
        for _, row in target_coords_df.iterrows():
            target_coord_lookup[str(row["Accession #"])] = (
                float(row["Latitude"]),
                float(row["Longitude"]),
            )
        results_df["Geo Dist"] = (
            results_df["Accession #"].apply(
                lambda acc: _compute_geodesic(
                    query_coords,
                    target_coord_lookup.get(str(acc)),
                )
            )
        )
    else:
        acc_match = query_coords_df[
            query_coords_df["Accession #"]
            == str(query_sample["Accession #"])
        ]
        query_coords: tuple[float, float] | None = (
            (
                float(acc_match.iloc[0]["Latitude"]),
                float(acc_match.iloc[0]["Longitude"]),
            )
            if not acc_match.empty
            else None
        )
        target_coord_lookup = {}
        for _, row in target_coords_df.iterrows():
            target_coord_lookup[str(row["Site"])] = (
                float(row["Latitude"]),
                float(row["Longitude"]),
            )
        results_df["Geo Dist"] = results_df["Site"].apply(
            lambda site: _compute_geodesic(
                query_coords,
                target_coord_lookup.get(site),
            )
        )

    if direction == "artefact_to_geology":
        results_df = results_df.rename(columns={
            "Accession #": "Geo Acc #",
            "Site": "Geo Site",
        })
    else:
        results_df = results_df.rename(columns={
            "Accession #": "Artefact Acc #",
            "Site": "Artefact Site",
            "Region": "Artefact Region",
        })

    return results_df


# -----------------------------
# Table Styling Functions
# -----------------------------
def highlight_geo_dist(s: pd.Series) -> list[str]:
    """Apply background color to Geo Dist cells."""
    styles = []
    for val in s:
        try:
            num = float(str(val).replace(" km", ""))
        except (ValueError, TypeError, AttributeError):
            styles.append("")
            continue
        if num < GEO_VERY_CLOSE_KM:
            color = COLORS["very_strong"]
        elif num < GEO_CLOSE_KM:
            color = COLORS["strong"]
        elif num < GEO_MODERATE_KM:
            color = COLORS["moderate"]
        else:
            color = COLORS["weak"]
        styles.append(
            f"background-color: {color}; {CELL_STYLE};"
        )
    return styles


def _color_dist_cell(
    val: Any,  # noqa: ANN401
    thresholds: tuple[float, float, float],
) -> str:
    """Apply background color to a distance cell."""
    try:
        v = float(val)
    except (ValueError, TypeError):
        return ""
    t1, t2, t3 = thresholds
    if v < t1:
        color = COLORS["very_strong"]
    elif v < t2:
        color = COLORS["strong"]
    elif v < t3:
        color = COLORS["moderate"]
    else:
        color = COLORS["weak"]
    return (
        f"background-color: {color}; {CELL_STYLE};"
        " border-radius: 0px"
    )


def color_eucl_dist(val: Any) -> str:  # noqa: ANN401
    """Apply background color to Eucl Dist cells."""
    return _color_dist_cell(
        val, (EUCL_VERY_STRONG, EUCL_STRONG, EUCL_MODERATE),
    )


def _color_aitch_with_thresholds(
    val: Any,  # noqa: ANN401
    thresholds: tuple[float, float, float] = (
        AITCH_VERY_STRONG, AITCH_STRONG, AITCH_MODERATE,
    ),
) -> str:
    """Apply background color using given thresholds."""
    return _color_dist_cell(val, thresholds)


# -----------------------------
# Excel Export with Formatting
# -----------------------------
def _get_fill_for_value(
    val: Any,  # noqa: ANN401
    thresholds: tuple[float, float, float],
) -> PatternFill | None:
    """Return a PatternFill for a numeric value."""
    try:
        v = float(val)
    except (ValueError, TypeError):
        return None
    t1, t2, t3 = thresholds
    if v < t1:
        return PatternFill(
            start_color="ADD8E6", fill_type="solid",
        )
    if v < t2:
        return PatternFill(
            start_color="90EE90", fill_type="solid",
        )
    if v < t3:
        return PatternFill(
            start_color="FFDAB9", fill_type="solid",
        )
    return PatternFill(
        start_color="F08080", fill_type="solid",
    )


def _get_geo_fill(val: Any) -> PatternFill | None:  # noqa: ANN401
    """Return a PatternFill for a Geo Dist string value."""
    try:
        num = float(str(val).replace(" km", ""))
    except (ValueError, TypeError, AttributeError):
        return None
    return _get_fill_for_value(
        num,
        (GEO_VERY_CLOSE_KM, GEO_CLOSE_KM, GEO_MODERATE_KM),
    )


def _create_styled_excel(
    results_df: pd.DataFrame,
    query_accession: str,
    query_site: str,
    direction: str,
    mode_key: str = "trace",
) -> bytes:
    """Create a styled Excel workbook with color-coded cells."""
    output = BytesIO()
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:  # pyright: ignore[reportArgumentType]
        results_df.to_excel(
            writer, sheet_name="Results",
            index=True, startrow=2,
        )
        ws = writer.sheets["Results"]

        # Header metadata
        ws["A1"] = (
            f"Query: {query_accession} — {query_site}"
        )
        ws["A1"].font = Font(bold=True, size=12)
        dir_label = (
            "Artefact -> Geology"
            if direction == "artefact_to_geology"
            else "Geology -> Artefact"
        )
        ws["A2"] = (
            f"Direction: {dir_label} | "
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )

        # Find column letters for distance columns
        # row where pandas writes headers
        # (0-indexed startrow=2 -> Excel row 3)
        header_row = 3
        col_map: dict[str, int] = {}
        for cell in ws[header_row]:
            if cell.value in (
                "Aitch Dist", "Eucl Dist", "Geo Dist",
            ):
                col_map[cell.value] = cell.column

        aitch_thresholds = _aitch_thresholds(mode_key)
        eucl_thresholds = (
            EUCL_VERY_STRONG, EUCL_STRONG, EUCL_MODERATE,
        )

        # Apply fills to data rows
        for row in ws.iter_rows(
            min_row=header_row + 1, max_row=ws.max_row,
        ):
            for cell in row:
                cell.border = thin_border
                if cell.column == col_map.get("Aitch Dist"):
                    fill = _get_fill_for_value(
                        cell.value, aitch_thresholds,
                    )
                    if fill:
                        cell.fill = fill
                elif cell.column == col_map.get("Eucl Dist"):
                    fill = _get_fill_for_value(
                        cell.value, eucl_thresholds,
                    )
                    if fill:
                        cell.fill = fill
                elif cell.column == col_map.get("Geo Dist"):
                    fill = _get_geo_fill(cell.value)
                    if fill:
                        cell.fill = fill

        # Style header row
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.border = thin_border

    return output.getvalue()


# -----------------------------
# UI Display Functions
# -----------------------------
def _to_image_bytes(img_src: str | bytes) -> bytes | None:
    """Convert an image source (file path or bytes) to bytes."""
    if isinstance(img_src, bytes):
        return img_src
    try:
        with open(img_src, "rb") as f:
            return f.read()
    except Exception:
        return None


def _render_zoom_viewer(
    img_src: bytes | tuple[str, str],
    uid: str,
    height: int = 450,
) -> None:
    """Render an interactive pan/zoom viewer via HTML.

    img_src can be:
    - bytes: image data (base64-encoded into HTML)
    - tuple[str, str]: (thumbnail_url, full_res_url)
      for progressive loading — shows thumbnail first,
      then auto-swaps to full resolution for max detail.
    """
    cid, iid = f"zc_{uid}", f"zi_{uid}"

    if isinstance(img_src, tuple):
        thumb_url, full_url = img_src
        img_attr = f'src="{thumb_url}"'
        # Progressive: load full-res in background, swap when
        # ready while preserving the user's current view.
        progressive_js = (
            "var ow=0;"
            f"var hd=new Image();hd.src='{full_url}';"
            "hd.onload=function(){"
            "var nw=hd.naturalWidth;"
            "if(ow>0){s*=(ow/nw);}"
            "hdReady=true;"
            f"im.src='{full_url}';"
            "};"
        )
    else:
        b64 = base64.b64encode(img_src).decode("ascii")
        img_attr = f'src="data:image/jpeg;base64,{b64}"'
        progressive_js = ""

    html = (
        f'<div id="{cid}" style="width:100%;height:{height}px;'
        "overflow:hidden;position:relative;background:#1a1a1a;"
        'cursor:grab;border-radius:6px;">'
        f'<img id="{iid}" {img_attr} '
        'style="position:absolute;max-width:none;'
        "image-orientation:from-image;"
        'transform-origin:0 0;" draggable="false">'
        "</div>"
        "<script>(function(){"
        f"const c=document.getElementById('{cid}'),"
        f"im=document.getElementById('{iid}');"
        "let s=1,tx=0,ty=0,drag=0,sx,sy,hdReady=false;"
        "function up(){im.style.transform="
        "'translate('+tx+'px,'+ty+'px) scale('+s+')';}"
        "function fit(){"
        "let rx=c.clientWidth/im.naturalWidth,"
        "ry=c.clientHeight/im.naturalHeight;"
        "s=Math.min(rx,ry);"
        "tx=(c.clientWidth-im.naturalWidth*s)/2;"
        "ty=(c.clientHeight-im.naturalHeight*s)/2;up();}"
        "im.onload=function(){"
        "if(!hdReady){fit();ow=im.naturalWidth;}"
        "else{up();}};"
        + progressive_js +
        "c.onwheel=function(e){e.preventDefault();"
        "let r=c.getBoundingClientRect(),"
        "mx=e.clientX-r.left,my=e.clientY-r.top,"
        "ps=s;s*=e.deltaY<0?1.15:0.87;"
        "s=Math.max(0.1,Math.min(30,s));"
        "tx=mx-(mx-tx)*(s/ps);ty=my-(my-ty)*(s/ps);up();};"
        "c.onmousedown=function(e){"
        "drag=1;sx=e.clientX-tx;sy=e.clientY-ty;"
        "c.style.cursor='grabbing';};"
        "c.onmousemove=function(e){"
        "if(!drag)return;tx=e.clientX-sx;"
        "ty=e.clientY-sy;up();};"
        "c.onmouseup=c.onmouseleave=function(){"
        "drag=0;c.style.cursor='grab';};"
        "c.ondblclick=function(){fit();};"
        "})();</script>"
    )
    st.components.v1.html(html, height=height + 10)


def display_artefact_photos(
    accession: str,
    photo_df: pd.DataFrame,
    photos_base_dir: str = "",
    access_token: str = "",
    share_url: str = "",
) -> None:
    """Display artefact photographs from local files or OneDrive."""
    images = _get_artefact_images(
        accession, photo_df, photos_base_dir,
        access_token, share_url,
    )
    if not images:
        # Diagnostic: show why no images were found
        row = photo_df[
            photo_df["Accession #"] == str(accession).strip()
        ] if not photo_df.empty else pd.DataFrame()
        if row.empty:
            st.caption(
                f"No photo mapping for accession {accession}"
            )
        else:
            err = st.session_state.get(
                "_graph_last_error", "",
            )
            st.warning(
                f"Photo mapping found but images failed to "
                f"load. {f'Last error: {err}' if err else ''}"
            )
        return
    with st.expander(
        f"Artefact Photographs ({len(images)} images)"
        " — scroll to zoom, drag to pan, "
        "double-click to reset",
        expanded=True,
    ):
        cols = st.columns(min(len(images), 3))
        for i, img_src in enumerate(images):
            with cols[i % 3]:
                # URL tuples go straight to the viewer
                if isinstance(img_src, tuple):
                    _render_zoom_viewer(
                        img_src,
                        uid=f"{accession}_{i}",
                    )
                    st.caption(f"View {i + 1}")
                else:
                    img_bytes = _to_image_bytes(img_src)
                    if img_bytes:
                        _render_zoom_viewer(
                            img_bytes,
                            uid=f"{accession}_{i}",
                        )
                        st.caption(f"View {i + 1}")
                    else:
                        st.caption(
                            f"View {i + 1}: not available"
                        )


def display_legend(mode_key: str = "trace") -> None:
    """Display color-coded threshold explanations."""
    vs, s, m = _aitch_thresholds(mode_key)
    show_eucl = _has_euclidean(mode_key)

    if show_eucl:
        col1, col2, col3 = st.columns(3)
    else:
        col1, col3 = st.columns(2)

    aitch_label = (
        "**Aitchison Distance** (ALR-5)"
        if mode_key == "alr5"
        else "**Aitchison Distance** (compositional)"
    )
    with col1:
        st.markdown(aitch_label)
        st.markdown(
            f"- :blue[< {vs}] "
            "-- Very strong match\n"
            f"- :green[{vs} - "
            f"{s}] -- Strong match\n"
            f"- :orange[{s} - "
            f"{m}] -- Moderate match\n"
            f"- :red[> {m}] -- Weak match"
        )
    if show_eucl:
        with col2:
            st.markdown("**Euclidean Distance** (geometric)")
            st.markdown(
                f"- :blue[< {EUCL_VERY_STRONG}] "
                "-- Very strong match\n"
                f"- :green[{EUCL_VERY_STRONG} - "
                f"{EUCL_STRONG}] -- Strong match\n"
                f"- :orange[{EUCL_STRONG} - "
                f"{EUCL_MODERATE}] -- Moderate match\n"
                f"- :red[> {EUCL_MODERATE}] -- Weak match"
            )
    with col3:
        st.markdown("**Geographical Distance**")
        st.markdown(
            f"- :blue[< {GEO_VERY_CLOSE_KM} km] "
            "-- Very close\n"
            f"- :green[{GEO_VERY_CLOSE_KM} - "
            f"{GEO_CLOSE_KM} km] -- Close\n"
            f"- :orange[{GEO_CLOSE_KM} - "
            f"{GEO_MODERATE_KM} km] -- Moderate distance\n"
            f"- :red[> {GEO_MODERATE_KM} km] -- Distant"
        )


def display_results_table(
    results_df: pd.DataFrame,
    query_accession: str,
    direction: str,
    query_site: str = "",
    mode_key: str = "trace",
) -> None:
    """Style and display the results table with downloads."""
    fmt: dict[str, str] = {"Aitch Dist": "{:.2f}"}
    if _has_euclidean(mode_key):
        fmt["Eucl Dist"] = "{:.2f}"
    styled_df = results_df.style.format(fmt)

    aitch_styler = partial(
        _color_aitch_with_thresholds,
        thresholds=_aitch_thresholds(mode_key),
    )
    styled_df = styled_df.apply(  # pyright: ignore[reportAttributeAccessIssue]
        highlight_geo_dist, subset=["Geo Dist"],
    ).map(aitch_styler, subset=["Aitch Dist"])
    if _has_euclidean(mode_key):
        styled_df = styled_df.map(
            color_eucl_dist, subset=["Eucl Dist"],
        )
    border_style = ("border", "1px solid gray")
    padding = ("padding", "4px")
    no_radius = ("border-radius", "0px")
    styled_df = styled_df.set_table_styles([
        {
            "selector": "table",
            "props": [
                ("border-collapse", "collapse"),
                border_style,
            ],
        },
        {
            "selector": "th",
            "props": [border_style, padding, no_radius],
        },
        {
            "selector": "td",
            "props": [border_style, padding, no_radius],
        },
    ])
    st.table(styled_df)

    direction_label = (
        "art2geo"
        if direction == "artefact_to_geology"
        else "geo2art"
    )
    base_filename = (
        f"GSF_{query_accession}"
        f"_{direction_label}_top{len(results_df)}"
    )

    dl_col1, dl_col2 = st.columns(2)
    with dl_col1:
        csv_data = (
            results_df.to_csv(index=False).encode("utf-8")
        )
        st.download_button(
            label="Download as CSV",
            data=csv_data,
            file_name=f"{base_filename}.csv",
            mime="text/csv",
        )
    with dl_col2:
        excel_data = _create_styled_excel(
            results_df, query_accession,
            query_site, direction,
            mode_key=mode_key,
        )
        st.download_button(
            label="Download as Excel",
            data=excel_data,
            file_name=f"{base_filename}.xlsx",
            mime=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )


def display_scatter_plot(
    results_df: pd.DataFrame,
    mode_key: str = "trace",
) -> None:
    """Create a scatter plot of Geo Dist vs. Aitch Dist."""
    plot_df = results_df[
        results_df["Geo Dist"] != "Unknown"
    ].copy()
    if plot_df.empty:
        st.warning(
            "No rows with known geographical distance to plot."
        )
        return

    plot_df["Geo Dist Num"] = plot_df["Geo Dist"].apply(  # pyright: ignore[reportAttributeAccessIssue]
        lambda v: float(str(v).replace(" km", ""))
    )

    hover_cols: dict[str, bool] = {}
    for col in [
        "Geo Acc #", "Geo Site",
        "Artefact Acc #", "Artefact Site",
    ]:
        if col in plot_df.columns:
            hover_cols[col] = True

    color_col = None
    color_map = None
    if "Lithology" in plot_df.columns:
        color_col = "Lithology"
        color_map = LITHOLOGY_COLOR_MAP
    elif "Artefact Region" in plot_df.columns:
        color_col = "Artefact Region"
        color_map = REGION_COLOR_MAP

    fig = px.scatter(
        plot_df,
        x="Geo Dist Num",
        y="Aitch Dist",
        color=color_col,
        hover_data=hover_cols,
        title=(
            "Scatter Plot of Geographical Distance "
            "vs. Aitchison Distance"
        ),
        color_discrete_map=color_map if color_map else {},
    )

    for ref in [
        GEO_VERY_CLOSE_KM, GEO_CLOSE_KM, GEO_MODERATE_KM,
    ]:
        fig.add_vline(
            x=ref, line_dash="dash",
            line_color="black", opacity=0.5,
        )
    for ref in _aitch_thresholds(mode_key):
        fig.add_hline(
            y=ref, line_dash="dash",
            line_color="black", opacity=0.5,
        )

    fig.update_traces(
        marker={
            "size": 20,
            "line": {"color": "black", "width": 2},
            "opacity": 0.8,
        },
    )
    fig.update_layout(
        template=None,
        paper_bgcolor="white",
        plot_bgcolor="white",
        font={"color": "black"},
        xaxis_title="Geographical Distance (km)",
        yaxis_title="Aitchison Distance",
        margin={"l": 40, "r": 40, "t": 40, "b": 40},
        legend_title=color_col if color_col else "Legend",
    )
    for ax_update in (fig.update_xaxes, fig.update_yaxes):
        ax_update(
            rangemode="tozero",
            showline=True,
            linecolor="black",
            linewidth=1,
            mirror=True,
            ticks="outside",
            tickfont={"color": "black"},
            showgrid=False,
            zeroline=False,
        )
    st.plotly_chart(fig, use_container_width=True)



def _aitch_to_marker_color(
    val: float, mode_key: str = "trace",
) -> str:
    """Map Aitchison distance to a marker color."""
    _, _, mod = _aitch_thresholds(mode_key)
    # Scale the 0.5-step bands proportionally
    step = mod / 6.0  # 6 bands below moderate
    if val < step:
        return "lightskyblue"
    if val < 2 * step:
        return "steelblue"
    if val < 3 * step:
        return "limegreen"
    if val < 4 * step:
        return "gold"
    if val < 5 * step:
        return "orange"
    if val < mod:
        return "tomato"
    return "darkred"


def display_results_map(
    results_df: pd.DataFrame,
    query_coords: tuple[float, float] | None,
    target_coords_df: pd.DataFrame,
    query_label: str,
    direction: str,
    mode_key: str = "trace",
) -> None:
    """Display an interactive map with query and match locations."""
    acc_col = (
        "Geo Acc #"
        if direction == "artefact_to_geology"
        else "Artefact Acc #"
    )
    site_col = (
        "Geo Site"
        if direction == "artefact_to_geology"
        else "Artefact Site"
    )

    map_points: list[dict[str, Any]] = []
    if direction == "artefact_to_geology":
        for _, row in results_df.iterrows():
            if row["Geo Dist"] == "Unknown":
                continue
            acc = str(row[acc_col])
            coord_row = target_coords_df[
                target_coords_df["Accession #"].astype(str)
                == acc
            ]
            if not coord_row.empty:
                map_points.append({
                    "lat": coord_row.iloc[0]["Latitude"],
                    "lon": coord_row.iloc[0]["Longitude"],
                    "label": f"{acc} - {row[site_col]}",
                    "aitch": row["Aitch Dist"],
                })
    else:
        for _, row in results_df.iterrows():
            if row["Geo Dist"] == "Unknown":
                continue
            site = row[site_col]
            coord_row = target_coords_df[
                target_coords_df["Site"] == site
            ]
            if not coord_row.empty:
                map_points.append({
                    "lat": coord_row.iloc[0]["Latitude"],
                    "lon": coord_row.iloc[0]["Longitude"],
                    "label": f"{row[acc_col]} - {site}",
                    "aitch": row["Aitch Dist"],
                })

    if not map_points and query_coords is None:
        return

    fig = go.Figure()

    # Invisible anchors at map-extent corners so fitbounds always
    # shows the full southern-Iberia region regardless of data spread
    _MAP_BOUNDS = {
        "west":  -9.8,
        "east":   1.5,
        "south": 33.5,
        "north": 41.5,
    }
    fig.add_trace(go.Scattermapbox(
        lat=[_MAP_BOUNDS["south"], _MAP_BOUNDS["north"]],
        lon=[_MAP_BOUNDS["west"], _MAP_BOUNDS["east"]],
        mode="markers",
        marker={"size": 1, "opacity": 0},
        showlegend=False,
        hoverinfo="skip",
    ))

    # Match points colored by Aitchison distance
    if map_points:
        map_df = pd.DataFrame(map_points)
        map_df["marker_color"] = map_df["aitch"].apply(
            lambda v: _aitch_to_marker_color(v, mode_key),
        )
        map_df["hover_text"] = map_df.apply(
            lambda r: (
                f"{r['label']}<br>Aitch: {r['aitch']:.2f}"
            ),
            axis=1,
        )

        # Black shadow trace behind all match points — simulates borders
        # (Scattermapbox marker.line is not supported)
        fig.add_trace(go.Scattermapbox(
            lat=map_df["lat"].tolist(),
            lon=map_df["lon"].tolist(),
            mode="markers",
            marker={"size": 16, "color": "black"},
            showlegend=False,
            hoverinfo="skip",
        ))

        # One trace per color group for legend entries.
        # Add traces worst-first so that better matches render
        # on top when multiple samples share the same coordinates.
        # legendrank keeps the legend ordered best-first.
        _, _, mod = _aitch_thresholds(mode_key)
        step = mod / 6.0
        def _fmt(v: float) -> str:
            return f"{v:.2f}" if v != int(v) else f"{v:.1f}"
        color_labels = [
            ("lightskyblue", 1,
             f"< {_fmt(step)} (very strong)"),
            ("steelblue", 2,
             f"{_fmt(step)}–{_fmt(2*step)} (very strong)"),
            ("limegreen", 3,
             f"{_fmt(2*step)}–{_fmt(3*step)} (strong)"),
            ("gold", 4,
             f"{_fmt(3*step)}–{_fmt(4*step)} (strong)"),
            ("orange", 5,
             f"{_fmt(4*step)}–{_fmt(5*step)} (moderate)"),
            ("tomato", 6,
             f"{_fmt(5*step)}–{_fmt(mod)} (moderate)"),
            ("darkred", 7,
             f"> {_fmt(mod)} (weak)"),
        ]
        for color, rank, label in reversed(color_labels):
            subset = map_df[
                map_df["marker_color"] == color
            ]
            if subset.empty:
                continue
            fig.add_trace(go.Scattermapbox(
                lat=subset["lat"],
                lon=subset["lon"],
                mode="markers",
                marker={"size": 12, "color": color},
                name=label,
                text=subset["hover_text"],
                legendrank=rank,
            ))

    # Query point — added last so it always renders on top.
    # White halo then black fill so it is distinct from match markers.
    if query_coords:
        fig.add_trace(go.Scattermapbox(
            lat=[query_coords[0]],
            lon=[query_coords[1]],
            mode="markers",
            marker={"size": 21, "color": "white"},
            showlegend=False,
            hoverinfo="skip",
        ))
        fig.add_trace(go.Scattermapbox(
            lat=[query_coords[0]],
            lon=[query_coords[1]],
            mode="markers",
            marker={"size": 15, "color": "black"},
            name=query_label,
            text=[query_label],
            legendrank=0,
        ))

    # Scale bar — ArcGIS-style alternating bar 0-50-100-200 km
    # Bottom-left corner of map, Atlantic Ocean west of Strait of Gibraltar
    # At 35.1°N: 111.32 × cos(35.1°) ≈ 91.1 km/° lon
    _sb_lat    = 35.1
    _sb_lon0   = -9.7
    _sb_dlon   = 50 / 91.1          # ≈ 0.549° per 50 km
    _sb_lon50  = _sb_lon0 + _sb_dlon
    _sb_lon100 = _sb_lon0 + 2 * _sb_dlon
    _sb_lon200 = _sb_lon0 + 4 * _sb_dlon
    _sb_lw     = 10   # bar thickness in screen pixels

    # Black base bar (0 → 200 km)
    fig.add_trace(go.Scattermapbox(
        lat=[_sb_lat, _sb_lat], lon=[_sb_lon0, _sb_lon200],
        mode="lines", line={"width": _sb_lw, "color": "#000000"},
        showlegend=False, hoverinfo="skip",
    ))
    # White cutout for 50-100 km segment — slightly narrower to preserve black outline
    fig.add_trace(go.Scattermapbox(
        lat=[_sb_lat, _sb_lat], lon=[_sb_lon50, _sb_lon100],
        mode="lines", line={"width": _sb_lw - 4, "color": "#ffffff"},
        showlegend=False, hoverinfo="skip",
    ))
    # Distance labels — Scattermapbox text traces so they pan with the bar
    _sb_num_lat = _sb_lat   # numbers sit just above bar via textposition
    _sb_km_lat  = _sb_lat + 0.18  # "Kilometers" title above the numbers
    fig.add_trace(go.Scattermapbox(
        lat=[_sb_num_lat] * 4,
        lon=[_sb_lon0, _sb_lon50, _sb_lon100, _sb_lon200],
        mode="text",
        text=["0", "50", "100", "200"],
        textfont={"size": 11, "color": "black"},
        textposition="top center",
        showlegend=False,
        hoverinfo="skip",
    ))
    fig.add_trace(go.Scattermapbox(
        lat=[_sb_km_lat],
        lon=[(_sb_lon0 + _sb_lon200) / 2],
        mode="text",
        text=["Kilometers"],
        textfont={"size": 11, "color": "black"},
        textposition="top center",
        showlegend=False,
        hoverinfo="skip",
    ))

    # Map layout — fitbounds="locations" auto-zooms to show every trace
    # (including the invisible corner anchors) so the full region is
    # always visible regardless of container size or aspect ratio.
    fig.update_layout(
        mapbox={
            "style": "white-bg",
            "layers": [
                {
                    "below": "traces",
                    "sourcetype": "raster",
                    "source": [
                        "https://server.arcgisonline.com/ArcGIS/rest/"
                        "services/NatGeo_World_Map/MapServer/tile"
                        "/{z}/{y}/{x}"
                    ],
                    "sourceattribution": (
                        "Tiles © Esri — National Geographic, Esri,"
                        " DeLorme, NAVTEQ, USGS, NRCAN, GEBCO, NOAA"
                    ),
                }
            ],
            "bounds": _MAP_BOUNDS,
            "center": {"lat": 37.5, "lon": -4.15},  # fallback
            "zoom": 6,                               # fallback
        },
        margin={"l": 0, "r": 0, "t": 30, "b": 0},
        title="Geographical Distribution of Matches",
    )
    # Responsive height: fill available viewport, scoped to the mapbox chart only
    st.markdown(
        """
        <style>
        [data-testid="stPlotlyChart"]:has(.mapboxgl-canvas) {
            height: calc(100vh - 160px) !important;
            min-height: 450px;
        }
        [data-testid="stPlotlyChart"]:has(.mapboxgl-canvas) > div {
            height: 100% !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.plotly_chart(fig, use_container_width=True)


# -----------------------------
# Radar / Spider Chart
# -----------------------------
def display_radar_chart(
    query_sample: pd.Series,
    results_df: pd.DataFrame,
    target_df: pd.DataFrame,
    query_accession: str,
    direction: str,
    mode_key: str = "trace",
) -> None:
    """Display a radar chart comparing ratio profiles."""
    ratio_cols = _get_ratio_columns(query_sample)
    if len(ratio_cols) < 3:
        st.info("Not enough ratio columns for a radar chart.")
        return

    # Build options for match selection
    acc_col = (
        "Geo Acc #"
        if direction == "artefact_to_geology"
        else "Artefact Acc #"
    )
    site_col = (
        "Geo Site"
        if direction == "artefact_to_geology"
        else "Artefact Site"
    )

    match_options: list[str] = []
    for _, row in results_df.iterrows():
        match_options.append(
            f"{row[acc_col]} - {row[site_col]}"
        )

    if not match_options:
        return

    selected = st.multiselect(
        "Select matches to compare on radar chart:",
        match_options,
        default=match_options[:min(3, len(match_options))],
        max_selections=8,
        key="radar_select",
    )

    if not selected:
        return

    # Collect ratio values
    query_values = np.asarray(pd.Series(pd.to_numeric(
        query_sample[ratio_cols], errors="coerce",
    )).fillna(EPSILON), dtype=float)
    all_values = [query_values]
    labels = [f"Query: {query_accession}"]

    for sel in selected:
        acc = sel.split(" - ")[0]
        match_rows = target_df[
            target_df["Accession #"] == acc
        ]
        if match_rows.empty:
            continue
        match_sample = match_rows.iloc[0]
        match_vals = np.asarray(pd.Series(pd.to_numeric(
            match_sample[ratio_cols], errors="coerce",
        )).fillna(EPSILON), dtype=float)
        all_values.append(match_vals)
        labels.append(sel)

    if len(all_values) < 2:
        st.info("No valid match data found for radar chart.")
        return

    # Normalize: min-max across all compared samples
    matrix = np.array(all_values)
    if mode_key == "alr5":
        # ALR values are already log-scale; skip log10
        work = matrix
    else:
        work = np.log10(np.clip(matrix, EPSILON, None))
    col_min = work.min(axis=0)
    col_max = work.max(axis=0)
    col_range = col_max - col_min
    col_range[col_range == 0] = 1
    normalized = (work - col_min) / col_range

    # Build radar chart
    fig = go.Figure()
    theta = [*list(ratio_cols), ratio_cols[0]]

    chart_colors = [
        "red", "#1f77b4", "#ff7f0e", "#2ca02c",
        "#d62728", "#9467bd", "#8c564b", "#e377c2",
        "#7f7f7f",
    ]
    for i, (vals, label) in enumerate(
        zip(normalized, labels, strict=False)
    ):
        r = [*list(vals), vals[0]]
        fig.add_trace(go.Scatterpolar(
            r=r,
            theta=theta,
            name=label,
            line={
                "width": 3 if i == 0 else 2,
                "color": chart_colors[
                    i % len(chart_colors)
                ],
            },
            fill="none",
        ))

    fig.update_layout(
        polar={
            "radialaxis": {
                "visible": True, "range": [0, 1],
            },
        },
        showlegend=True,
        title="Ratio Profile Comparison (log-normalized)",
        margin={"l": 60, "r": 60, "t": 50, "b": 40},
    )
    st.plotly_chart(fig, use_container_width=True)


# -----------------------------
# Query Execution Helpers
# -----------------------------
def _get_query_coords(
    query_sample: pd.Series,
    coords_df: pd.DataFrame,
    direction: str,
) -> tuple[float, float] | None:
    """Look up the coordinates for a query sample."""
    if direction == "artefact_to_geology":
        match = coords_df[
            coords_df["Site"] == query_sample["Site"]
        ]
    else:
        match = coords_df[
            coords_df["Accession #"].astype(str)
            == str(query_sample["Accession #"])
        ]
    if not match.empty:
        return (
            float(match.iloc[0]["Latitude"]),
            float(match.iloc[0]["Longitude"]),
        )
    return None


def _resolve_sample(
    accession_number: str,
    source_df: pd.DataFrame,
    label_prefix: str,
    key_suffix: str = "",
) -> pd.Series | None:
    """Find a sample by accession number."""
    matches = source_df[
        source_df["Accession #"] == accession_number
    ]
    if matches.empty:
        st.error(
            f"No {label_prefix} found with "
            f"Accession # {accession_number}"
        )
        return None
    if len(matches) == 1:
        return matches.iloc[0]
    st.info(
        f"Multiple {label_prefix}s found for "
        f"Accession # {accession_number}. "
        "Please select a site:"
    )
    sites = list(matches["Site"].unique())  # pyright: ignore[reportAttributeAccessIssue]
    selected_site = str(st.selectbox(
        "Select a site",
        sites,
        key=f"site_select_{key_suffix}",
    ))
    site_mask = matches["Site"].str.lower() == selected_site.lower()  # pyright: ignore[reportAttributeAccessIssue]
    return matches[site_mask].iloc[0]  # pyright: ignore[reportAttributeAccessIssue]


def _execute_query(
    accession_number: str,
    query_mode: str,
    top_n: int,
    selected_regions: list[str],
    artefact_df: pd.DataFrame,
    geology_df: pd.DataFrame,
    geo_coords_df: pd.DataFrame,
    arch_coords_df: pd.DataFrame,
    photo_df: pd.DataFrame | None = None,
    photos_base_dir: str = "",
    access_token: str = "",
    share_url: str = "",
    mode_key: str = "trace",
) -> None:
    """Execute a single query and display results."""
    if query_mode == "Artefact \u2192 Geology":
        direction = "artefact_to_geology"
        source_df = artefact_df
        target_df = geology_df
        query_coords_df = arch_coords_df
        target_coords_df = geo_coords_df
        label_prefix = "artefact"
        region_col = "Region"
    else:
        direction = "geology_to_artefact"
        source_df = geology_df
        target_df = artefact_df
        query_coords_df = geo_coords_df
        target_coords_df = arch_coords_df
        label_prefix = "geology sample"
        region_col = "Artefact Region"

    sample = _resolve_sample(
        accession_number, source_df,
        label_prefix, key_suffix="single",
    )
    if sample is None:
        return

    # Build header
    if "Region" in sample.index and bool(
        pd.notnull(sample.get("Region"))
    ):
        site_info = (
            f"{sample['Site']} ({sample['Region']})"
        )
    else:
        site_info = str(sample.get("Site", "Unknown Site"))
    st.subheader(
        f"Results for Accession # "
        f"{sample['Accession #']} - {site_info}"
    )

    # Data quality indicator
    total_ratios = _count_total_ratio_columns(source_df)
    valid_ratios = len(_get_ratio_columns(sample))
    if valid_ratios < total_ratios:
        st.warning(
            f"Data Quality: {valid_ratios}/{total_ratios}"
            " ratio columns available. "
            f"Missing {total_ratios - valid_ratios} "
            "values - results based on fewer dimensions."
        )
    else:
        st.info(
            f"Data Quality: {valid_ratios}/{total_ratios}"
            " ratio columns available (complete)."
        )

    # Get top matches
    results_df = get_top_matches(
        sample, target_df, query_coords_df,
        target_coords_df, top_n, direction,
        mode_key=mode_key,
    )

    # Apply region filter
    if (
        selected_regions
        and not results_df.empty
        and region_col in results_df.columns
    ):
        results_df = pd.DataFrame(
            results_df[
                results_df[region_col].isin(selected_regions)
            ]
        )

    if results_df.empty:
        st.warning(
            "No matching results found. "
            "Try adjusting the region filter "
            "or increasing the number of results."
        )
        return

    # Display results
    with st.expander(
        "Distance Interpretation Guide", expanded=False,
    ):
        display_legend(mode_key=mode_key)

    display_results_table(
        results_df, accession_number,
        direction, query_site=site_info,
        mode_key=mode_key,
    )

    # Display artefact photographs (if any source configured)
    if (
        (photos_base_dir or (access_token and share_url))
        and photo_df is not None
        and not photo_df.empty
        and direction == "artefact_to_geology"
    ):
        display_artefact_photos(
            str(sample["Accession #"]),
            photo_df,
            photos_base_dir,
            access_token=access_token,
            share_url=share_url,
        )

    display_radar_chart(
        sample, results_df, target_df,
        accession_number, direction,
        mode_key=mode_key,
    )
    display_scatter_plot(
        results_df, mode_key=mode_key,
    )

    query_coords = _get_query_coords(
        sample, query_coords_df, direction,
    )
    query_label = (
        f"Query: {sample['Accession #']} "
        f"- {sample['Site']}"
    )
    display_results_map(
        results_df, query_coords,
        target_coords_df, query_label, direction,
        mode_key=mode_key,
    )


# -----------------------------
# Batch Query
# -----------------------------
def _execute_batch_query(
    artefact_df: pd.DataFrame,
    geology_df: pd.DataFrame,
    geo_coords_df: pd.DataFrame,
    arch_coords_df: pd.DataFrame,
    mode_key: str = "trace",
) -> None:
    """Batch query: upload CSV, run matching for each."""
    uploaded_file = st.file_uploader(
        "Upload a CSV with accession numbers",
        type=["csv"],
        key="batch_upload",
    )
    if uploaded_file is None:
        st.info(
            "Upload a CSV file containing a column "
            "of accession numbers."
        )
        return

    input_df = pd.read_csv(uploaded_file)
    if input_df.empty:
        st.error("The uploaded CSV is empty.")
        return

    acc_column = st.selectbox(
        "Select the column containing accession numbers:",
        input_df.columns,
        key="batch_col",
    )
    query_mode = st.radio("Query Direction:", [
        "Artefact \u2192 Geology",
        "Geology \u2192 Artefact",
    ], key="batch_mode")
    top_n = st.number_input(
        "Top N per query:",
        min_value=1, max_value=MAX_TOP_N,
        value=DEFAULT_TOP_N, step=1,
        key="batch_topn",
    )

    if not st.button("Run Batch", key="batch_run"):
        return

    if query_mode == "Artefact \u2192 Geology":
        direction = "artefact_to_geology"
        source_df = artefact_df
        target_df = geology_df
        query_coords_df = arch_coords_df
        target_coords_df = geo_coords_df
    else:
        direction = "geology_to_artefact"
        source_df = geology_df
        target_df = artefact_df
        query_coords_df = geo_coords_df
        target_coords_df = arch_coords_df

    accession_numbers = (
        input_df[acc_column]
        .astype(str).str.strip()
        .dropna().unique().tolist()
    )
    if not accession_numbers:
        st.error(
            "No valid accession numbers found "
            "in the selected column."
        )
        return

    all_results: list[pd.DataFrame] = []
    errors: list[str] = []
    progress = st.progress(0, text="Processing batch...")
    total = len(accession_numbers)

    for i, acc in enumerate(accession_numbers):
        progress.progress(
            (i + 1) / total,
            text=f"Processing {acc} ({i + 1}/{total})",
        )
        matches = source_df[
            source_df["Accession #"] == acc
        ]
        if matches.empty:
            errors.append(acc)
            continue
        sample = matches.iloc[0]
        result = get_top_matches(
            sample, target_df, query_coords_df,
            target_coords_df, top_n, direction,
            mode_key=mode_key,
        )
        if not result.empty:
            result.insert(0, "Query Acc #", acc)
            all_results.append(result)

    progress.empty()

    if errors:
        st.warning(
            f"{len(errors)} accession number(s) not found: "
            f"{', '.join(errors[:20])}"
        )

    if not all_results:
        st.error(
            "No results found for any accession number "
            "in the batch."
        )
        return

    combined_df = pd.concat(all_results, ignore_index=True)
    combined_df.index += 1

    # Frequency summary
    acc_col = (
        "Geo Acc #"
        if direction == "artefact_to_geology"
        else "Artefact Acc #"
    )
    site_col = (
        "Geo Site"
        if direction == "artefact_to_geology"
        else "Artefact Site"
    )

    st.subheader("Source Frequency Summary")
    st.markdown(
        "Geological sources ranked by how often they "
        "appear across all queried samples:"
    )

    freq_df = (
        combined_df.groupby([acc_col, site_col])
        .agg(
            Count=("Query Acc #", "nunique"),
            Avg_Aitch=("Aitch Dist", "mean"),
            Min_Aitch=("Aitch Dist", "min"),
        )
        .sort_values(  # pyright: ignore[reportCallIssue]
            ["Count", "Avg_Aitch"],
            ascending=[False, True],
        )
        .reset_index()
    )
    freq_df["Avg_Aitch"] = np.round(
        freq_df["Avg_Aitch"], 2,
    )
    freq_df["Min_Aitch"] = np.round(
        freq_df["Min_Aitch"], 2,
    )
    aitch_styler = partial(
        _color_aitch_with_thresholds,
        thresholds=_aitch_thresholds(mode_key),
    )
    styled_freq = (
        freq_df.style
        .map(aitch_styler, subset=["Avg_Aitch", "Min_Aitch"])
        .format({"Avg_Aitch": "{:.2f}", "Min_Aitch": "{:.2f}"})
    )
    st.dataframe(styled_freq, use_container_width=True)

    # Combined results
    with st.expander("All Individual Results", expanded=False):
        st.dataframe(combined_df, use_container_width=True)

    # Downloads
    base_filename = (
        f"GSF_batch_{len(accession_numbers)}queries"
    )
    dl1, dl2 = st.columns(2)
    with dl1:
        st.download_button(
            "Download combined results as CSV",
            combined_df.to_csv(index=False).encode("utf-8"),
            file_name=f"{base_filename}.csv",
            mime="text/csv",
            key="batch_csv",
        )
    with dl2:
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:  # pyright: ignore[reportArgumentType]
            freq_df.to_excel(
                writer,
                sheet_name="Frequency Summary",
                index=False,
            )
            combined_df.to_excel(
                writer,
                sheet_name="All Results",
                index=False,
            )
        xl_mime = (
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
        st.download_button(
            "Download as Excel",
            output.getvalue(),
            file_name=f"{base_filename}.xlsx",
            mime=xl_mime,
            key="batch_xlsx",
        )

    n_success = len(accession_numbers) - len(errors)
    st.success(
        f"Batch complete: {n_success} successful queries, "
        f"{len(all_results)} results total."
    )


# -----------------------------
# Comparison Mode
# -----------------------------
def _execute_comparison(
    artefact_df: pd.DataFrame,
    geology_df: pd.DataFrame,
    geo_coords_df: pd.DataFrame,
    arch_coords_df: pd.DataFrame,
    photo_df: pd.DataFrame | None = None,
    photos_base_dir: str = "",
    access_token: str = "",
    share_url: str = "",
    mode_key: str = "trace",
) -> None:
    """Enter 2+ accession numbers to find shared sources."""
    st.markdown(
        "Enter 2 or more accession numbers to find "
        "**shared geological sources** that match "
        "multiple artefacts."
    )

    query_mode = st.radio("Query Direction:", [
        "Artefact \u2192 Geology",
        "Geology \u2192 Artefact",
    ], key="comp_mode")
    top_n = st.number_input(
        "Top N per sample:",
        min_value=1, max_value=MAX_TOP_N,
        value=DEFAULT_TOP_N, step=1,
        key="comp_topn",
    )

    acc_text = st.text_area(
        "Enter accession numbers "
        "(comma or newline separated):",
        height=100,
        key="comp_input",
        placeholder="e.g. 3003, 3004, 3005",
    )

    if not st.button("Compare", key="comp_run"):
        return

    # Parse accession numbers
    raw = acc_text.replace(",", "\n").strip()
    accession_numbers = [
        a.strip() for a in raw.split("\n") if a.strip()
    ]

    if len(accession_numbers) < 2:
        st.warning(
            "Please enter at least 2 accession numbers "
            "to compare."
        )
        return

    if query_mode == "Artefact \u2192 Geology":
        direction = "artefact_to_geology"
        source_df = artefact_df
        target_df = geology_df
        query_coords_df = arch_coords_df
        target_coords_df = geo_coords_df
    else:
        direction = "geology_to_artefact"
        source_df = geology_df
        target_df = artefact_df
        query_coords_df = geo_coords_df
        target_coords_df = arch_coords_df

    acc_col = (
        "Geo Acc #"
        if direction == "artefact_to_geology"
        else "Artefact Acc #"
    )
    site_col = (
        "Geo Site"
        if direction == "artefact_to_geology"
        else "Artefact Site"
    )

    # Run queries
    all_results: dict[str, pd.DataFrame] = {}
    errors: list[str] = []
    for acc in accession_numbers:
        matches = source_df[
            source_df["Accession #"] == acc
        ]
        if matches.empty:
            errors.append(acc)
            continue
        sample = matches.iloc[0]
        result = get_top_matches(
            sample, target_df, query_coords_df,
            target_coords_df, top_n, direction,
            mode_key=mode_key,
        )
        if not result.empty:
            all_results[acc] = result

    if errors:
        st.warning(f"Not found: {', '.join(errors)}")

    if len(all_results) < 2:
        st.error(
            "Need results from at least 2 accession "
            "numbers to compare."
        )
        return

    # Display photos of compared artefacts
    if (
        (photos_base_dir or (access_token and share_url))
        and photo_df is not None
        and not photo_df.empty
        and direction == "artefact_to_geology"
    ):
        photo_cols = st.columns(
            min(len(all_results), 4)
        )
        for i, acc in enumerate(all_results):
            with photo_cols[i % min(len(all_results), 4)]:
                images = _get_artefact_images(
                    acc, photo_df, photos_base_dir,
                    access_token, share_url,
                )
                if images:
                    try:
                        st.image(
                            images[0],
                            caption=f"Acc # {acc}",
                            use_container_width=True,
                        )
                    except Exception:
                        st.caption(f"Acc # {acc}")

    # Find shared sources
    source_counter: Counter[tuple[str, str]] = Counter()
    source_aitch: dict[
        tuple[str, str], dict[str, float]
    ] = {}

    for query_acc, result_df in all_results.items():
        for _, row in result_df.iterrows():
            source_key = (
                str(row[acc_col]), str(row[site_col]),
            )
            source_counter[source_key] += 1
            if source_key not in source_aitch:
                source_aitch[source_key] = {}
            source_aitch[source_key][query_acc] = float(
                row["Aitch Dist"]
            )

    # Build comparison table
    shared_sources = {
        k: v for k, v in source_counter.items() if v >= 2
    }

    if not shared_sources:
        st.info(
            "No shared geological sources found across "
            "the queried samples in their top-N results. "
            "Try increasing the number of results."
        )
        # Still show individual results
        for acc, result_df in all_results.items():
            with st.expander(
                f"Results for {acc}", expanded=False,
            ):
                st.dataframe(
                    result_df, use_container_width=True,
                )
        return

    st.subheader(
        f"Shared Sources ({len(shared_sources)} found)"
    )
    st.markdown(
        "Sources appearing in the top-N results of "
        "2 or more queried samples:"
    )

    query_accs = list(all_results.keys())
    comp_rows: list[dict[str, Any]] = []
    for (src_acc, src_site), count in sorted(
        shared_sources.items(),
        key=lambda x: (-x[1], x[0]),
    ):
        row_data: dict[str, Any] = {
            acc_col: src_acc,
            site_col: src_site,
            "Shared Count": count,
        }
        aitch_values = source_aitch.get(
            (src_acc, src_site), {},
        )
        for q_acc in query_accs:
            row_data[f"Aitch ({q_acc})"] = (
                aitch_values.get(q_acc, None)
            )
        # Add lithology if available
        if direction == "artefact_to_geology":
            for result_df in all_results.values():
                lit_row = result_df[
                    result_df[acc_col] == src_acc
                ]
                if (
                    not lit_row.empty
                    and "Lithology" in lit_row.columns
                ):
                    row_data["Lithology"] = (
                        lit_row.iloc[0]["Lithology"]
                    )
                    break
        comp_rows.append(row_data)

    comp_df = pd.DataFrame(comp_rows)

    # Sort by shared count desc, then average Aitchison dist
    aitch_cols = [
        c for c in comp_df.columns
        if c.startswith("Aitch (")
    ]
    if aitch_cols:
        comp_df["Avg Aitch"] = np.round(
            comp_df[aitch_cols].mean(axis=1), 2,
        )
        comp_df = comp_df.sort_values(  # pyright: ignore[reportCallIssue]
            ["Shared Count", "Avg Aitch"],
            ascending=[False, True],
        )
    comp_df.reset_index(drop=True, inplace=True)
    comp_df.index += 1

    # Highlight rows matching ALL queried samples
    def _highlight_shared(
        row: pd.Series,
    ) -> list[str]:
        if row["Shared Count"] == len(all_results):
            return [
                "background-color: #D4EDDA"
            ] * len(row)
        return [""] * len(row)

    styled = comp_df.style.apply(
        _highlight_shared, axis=1,
    )
    avg_cols = (
        ["Avg Aitch"]
        if "Avg Aitch" in comp_df.columns
        else []
    )
    aitch_color_fn = partial(
        _color_aitch_with_thresholds,
        thresholds=_aitch_thresholds(mode_key),
    )
    for col in aitch_cols + avg_cols:
        styled = styled.map(
            aitch_color_fn, subset=[col],
        )

    st.table(styled)
    st.caption(
        "Green rows = sources matching ALL queried samples."
    )

    # Individual results in expanders
    for acc, result_df in all_results.items():
        with st.expander(
            f"Full results for {acc}", expanded=False,
        ):
            st.dataframe(
                result_df, use_container_width=True,
            )

    # Download
    csv_data = comp_df.to_csv(index=False).encode("utf-8")
    st.download_button(
        "Download comparison as CSV",
        csv_data,
        file_name=(
            f"GSF_comparison_{len(query_accs)}samples.csv"
        ),
        mime="text/csv",
        key="comp_csv",
    )


# -----------------------------
# Main Application
# -----------------------------
def main() -> None:
    """Application entry point."""
    # Render sidebar to get settings (element mode)
    with st.sidebar:
        st.header("About")
        st.markdown(
            "**GSF - Geology Source Finder** matches "
            "archaeological artefacts to potential "
            "geological sources using trace element "
            "ratio comparison."
        )
        st.divider()
        st.header("Settings")
        element_mode = st.radio(
            "Element Mode:",
            [
                "Trace Elements (16 ratios)",
                "All Elements (22 ratios)",
                "ALR-5 Elements (5 log-ratios)",
            ],
            help=(
                "Trace Elements uses 16 trace-element "
                "ratios. All Elements adds 6 "
                "major-element ratios (22 total). "
                "ALR-5 uses 5 additive log-ratio "
                "coordinates from 6 key elements "
                "(Ni, Cr, Y, Nb, Sr, Zr)."
            ),
        )
        if "ALR" in element_mode:
            mode_key = "alr5"
        elif "Trace" in element_mode:
            mode_key = "trace"
        else:
            mode_key = "all"
        if mode_key == "alr5":
            st.caption(
                "**Metrics:** Aitchison Distance "
                "(ALR coordinates) · "
                "Geographical Distance"
            )
        else:
            st.caption(
                "**Metrics:** Aitchison Distance "
                "(CLR transform) · "
                "Euclidean Distance · "
                "Geographical Distance"
            )
        st.divider()
        saved_config = _load_config()

        # --- Local photos folder ---
        saved_dir = saved_config.get("photos_base_dir", "")
        photos_base_dir = st.text_input(
            "Local Photos Folder:",
            value=saved_dir,
            help=(
                "Local path to the folder containing "
                "artefact photographs (e.g. a OneDrive "
                "sync folder)."
            ),
            key="photos_dir",
        )
        if photos_base_dir != saved_dir:
            _save_config(
                {**saved_config, "photos_base_dir": photos_base_dir}
            )

        # --- OneDrive (Graph API) ---
        st.markdown("---")
        st.markdown("**OneDrive (online)**")

        saved_share = saved_config.get(
            "onedrive_folder_url", "",
        )
        share_url = st.text_input(
            "OneDrive Shared Folder URL:",
            value=saved_share,
            help=(
                "The 1drv.ms sharing link to the root "
                "photos folder in OneDrive."
            ),
            key="onedrive_url",
        )
        if share_url != saved_share:
            _save_config(
                {**saved_config, "onedrive_folder_url": share_url}
            )

        saved_client = saved_config.get("azure_client_id", "")
        client_id = st.text_input(
            "Azure App Client ID:",
            value=saved_client,
            help=(
                "Register a free Azure AD app at "
                "portal.azure.com, enable public client "
                "flows, add Files.Read permission, then "
                "paste the Application (client) ID here."
            ),
            key="azure_client_id",
            type="password",
        )
        if client_id != saved_client:
            _save_config(
                {**saved_config, "azure_client_id": client_id}
            )

        # Auth status and sign-in
        access_token = ""
        if client_id and share_url:
            access_token = (
                _get_valid_access_token(client_id) or ""
            )
            if access_token:
                st.success("OneDrive: signed in")
                if st.button(
                    "Sign out", key="graph_signout",
                ):
                    st.session_state.pop(
                        "graph_token_data", None,
                    )
                    path = _get_token_cache_path()
                    if os.path.exists(path):
                        os.remove(path)
                    st.rerun()
            else:
                dc = st.session_state.get("device_code_flow")
                if dc:
                    st.info(
                        f"Go to **{dc['verification_uri']}** "
                        f"and enter code: "
                        f"**{dc['user_code']}**"
                    )
                    if time.time() > dc.get("expires_at", 0):
                        st.error("Code expired. Try again.")
                        st.session_state.pop(
                            "device_code_flow", None,
                        )
                    elif st.button(
                        "Check sign-in status",
                        key="graph_poll",
                    ):
                        try:
                            result = _poll_for_token(
                                client_id, dc["device_code"],
                            )
                        except ValueError as exc:
                            st.error(str(exc))
                            st.session_state.pop(
                                "device_code_flow", None,
                            )
                            result = None
                        if result:
                            result["expires_at"] = (
                                time.time()
                                + result.get("expires_in", 3600)
                            )
                            st.session_state[
                                "graph_token_data"
                            ] = result
                            st.session_state.pop(
                                "device_code_flow", None,
                            )
                            _save_token_cache(result)
                            st.rerun()
                        else:
                            st.warning(
                                "Not ready yet. Complete "
                                "sign-in in browser, then "
                                "check again."
                            )
                else:
                    if st.button(
                        "Sign in to OneDrive",
                        key="graph_signin",
                    ):
                        flow = _start_device_code_flow(
                            client_id,
                        )
                        if flow:
                            flow["expires_at"] = (
                                time.time()
                                + flow.get("expires_in", 900)
                            )
                            st.session_state[
                                "device_code_flow"
                            ] = flow
                            st.rerun()
                        else:
                            st.error(
                                "Could not start sign-in. "
                                "Check the Client ID."
                            )

    st.title("GSF - Geology Source Finder")

    # Load data with selected element mode
    try:
        (
            artefact_df, geology_df,
            geo_coords_df, arch_coords_df,
        ) = load_data(mode_key)
    except FileNotFoundError as e:
        st.error(str(e))
        return
    except Exception as e:
        st.error("Error loading data.")
        st.exception(e)
        return

    # Load photo mapping
    photo_df = _load_photo_mapping()

    # Show dataset counts in sidebar
    with st.sidebar:
        st.divider()
        st.caption(
            f"Dataset: {len(artefact_df)} artefacts, "
            f"{len(geology_df)} geology samples"
        )
        if not photo_df.empty:
            st.caption(
                f"Photo mapping: {len(photo_df)} artefacts"
            )

    # Tab layout
    tab_single, tab_batch, tab_compare = st.tabs([
        "Single Query",
        "Batch Query",
        "Comparison",
    ])

    # --- Tab 1: Single Query ---
    with tab_single:
        query_mode = st.radio("Select Query Mode:", [
            "Artefact \u2192 Geology",
            "Geology \u2192 Artefact",
        ], key="single_mode")

        # Autocomplete accession number selection
        options_df = (
            artefact_df
            if query_mode == "Artefact \u2192 Geology"
            else geology_df
        )

        options = sorted(
            options_df.apply(
                lambda r: (
                    f"{r['Accession #']} - {r['Site']}"
                ),
                axis=1,
            ).unique()
        )

        selected = st.selectbox(
            "Select an accession number:",
            options=["", *options],
            format_func=lambda x: (
                "Type to search..." if x == "" else x
            ),
            key="single_acc",
        )
        accession_number = (
            selected.split(" - ")[0].strip()
            if selected
            else ""
        )

        top_n = st.number_input(
            "Number of top results:",
            min_value=1,
            max_value=MAX_TOP_N,
            value=DEFAULT_TOP_N,
            step=1,
            key="single_topn",
        )

        # Region filter
        all_regions = sorted(
            set(artefact_df["Region"].dropna().unique())
            | set(geology_df["Region"].dropna().unique())
        )
        selected_regions = st.multiselect(
            "Filter by region(s) (optional):",
            all_regions,
            key="single_regions",
        )

        # Execute query
        if accession_number:
            _execute_query(
                accession_number, query_mode,
                top_n, selected_regions,
                artefact_df, geology_df,
                geo_coords_df, arch_coords_df,
                photo_df=photo_df,
                photos_base_dir=photos_base_dir,
                access_token=access_token,
                share_url=share_url,
                mode_key=mode_key,
            )

    # --- Tab 2: Batch Query ---
    with tab_batch:
        _execute_batch_query(
            artefact_df, geology_df,
            geo_coords_df, arch_coords_df,
            mode_key=mode_key,
        )

    # --- Tab 3: Comparison Mode ---
    with tab_compare:
        _execute_comparison(
            artefact_df, geology_df,
            geo_coords_df, arch_coords_df,
            photo_df=photo_df,
            photos_base_dir=photos_base_dir,
            access_token=access_token,
            share_url=share_url,
            mode_key=mode_key,
        )


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        st.error("An unexpected error occurred.")
        st.exception(e)
